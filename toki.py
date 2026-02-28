import sys
import time
import pygame
from openpyxl import load_workbook
from datetime import date

def focusTime(focus):
    print("""
　　　　　 ／＞　　フ 
　　　　　 | 　_　_|
　 　　　／` ミ＿xノ  
　　 　 /　　　   |
　　　 /　 ヽ　　 ﾉ
      │　   |　|　|
　／￣|　　 |　|　|
　| (￣ヽ＿_ヽ_)__)
　＼二つ
    """)

    print("""
   __                     
  / _|                    
 | |_ ___   ___ _   _ ___ 
 |  _/ _ \\ / __| | | / __|
 | || (_) | (__| |_| \\__ \\
 |_| \\___/ \\___|\\__,_|___/
        """)

    while focus:
        mins, secs = divmod(focus, 60)
        timer = '{:02d}:{:02d}'.format(mins, secs)
        print(timer, end='\r')
        time.sleep(1)
        focus -= 1


def clear(r):
    for i in range(r):
        sys.stdout.write("\033[1A\033[K")
        sys.stdout.flush()

def restTime(rest):
    print("""
          
                _   
  _ __ ___  ___| |_ 
 | '__/ _ \\/ __| __|
 | | |  __/\\__ \\ |_ 
 |_|  \\___||___/\\__|
        """)
    while rest:
        mins, secs = divmod(rest, 60)
        timer = '{:02d}:{:02d}'.format(mins,secs)
        print(timer, end='\r')
        time.sleep(1)
        rest -= 1

def sound():
    pygame.mixer.init()
    mysound = pygame.mixer.Sound("Sound/finishNoise.mp3")
    mysound.play()
    mysound.set_volume(0.5)
    
def jazz(music):
    if music == "Y" or music == "y":
        pygame.mixer.init()
        jazzy = pygame.mixer.Sound("Sound/jazz.mp3")
        jazzy.play(-1)
        pygame.mixer.music.set_volume(0.7)

def rainSound(rain):
    if rain == "Y" or rain == "y":
        pygame.mixer.init()
        rainSound = pygame.mixer.Sound("Sound/rain.mp3")
        rainSound.play(-1)
        pygame.mixer.music.set_volume(0.7)


def log(date, cycles, focusTime):
    wb = load_workbook("log.xlsx")
    ws = wb.active

    Row = ws.max_row
    Row += 1
    ws.cell(row=Row, column=1).value = date
    ws.cell(row=Row, column=2).value = cycles
    ws.cell(row=Row, column =3).value = focusTime

    wb.save("log.xlsx")
    
def timeConverter(timeConvert):
    return time.strftime("%H:%M:%S", time.gmtime(timeConvert))


print("""\
  _______ ____  _  _______ 
 |__   __/ __ \\| |/ /_   _|
    | | | |  | | ' /  | |  
    | | | |  | |  <   | |  
    | | | |__| | . \\ _| |_ 
    |_|  \\____/|_|\\_\\_____|                          
""")

count = 0
focus = int(input("Enter focus time (in second): "))
rest = int(input("Enter resting time: "))
music = input("Do you want jazz music? (Y/n): ")
rain = input("Do you want rain ambience noise? (Y/n): ")


confirmation = input("Press y to begin:")
if confirmation != "y":
    print("Cancelling...")
    time.sleep(1)
    sys.exit(0)
else:
    clear(3)

jazz(music)
rainSound(rain)

while True:
    focusTime(focus)
    clear(8)
    sound()
    restTime(rest)
    count += 1
    confirm = input("Continue? (Y/N)")
    sys.stdout.flush()
    if confirm == "N" or confirm == "n":
        print(f"you completed {count} cycles!")
        time.sleep(1)
        print(f"{focus} second focus time and {rest} second rest time per cycle")
        break
    elif confirm == "Y" or confirm == "y":
        clear(20)

timeConvert = count * focus
focusLog = timeConverter(timeConvert)
cycleLog = count
dateLog = date.today()

log(dateLog, cycleLog, focusLog)

